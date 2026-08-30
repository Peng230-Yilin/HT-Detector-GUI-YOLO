# This Python file uses the following encoding: utf-8

import sys
import os
import io
import math
import numbers
import shutil
import tempfile
from pathlib import Path
from PySide6.QtWidgets import (QMainWindow, QFileDialog,QColorDialog, QComboBox,
                                QDialog, QFontComboBox,QTextEdit, QInputDialog,
                                QLineEdit, QMenu, QMessageBox,QProgressBar, QToolBar,
                                QVBoxLayout, QWidget, QTreeView, QTableView, QFileSystemModel,
                                QHeaderView, QHBoxLayout, QSplitter)
from PySide6.QtGui import (QAction, QGuiApplication, QIcon, QKeySequence, QStandardItemModel,
                            QStandardItem, QImage, QPixmap)
from PySide6.QtCore import (QUrl, Qt, Slot, Signal, QDir, QEvent, QThread,
                            QCoreApplication)

from PySide6.QtPrintSupport import (QAbstractPrintDialog, QPrinter,
                                    QPrintDialog, QPrintPreviewDialog)
from PySide6.QtSql import QSqlTableModel
from PySide6.QtGui import QIcon
#from tr import tr
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.axes import Axes
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg, NavigationToolbar2QT
from matplotlib.figure import Figure
from scipy import stats
import openpyxl

from ui import ui_detectmain
from camera import Camera
from interface_config import load_effective_settings
from yolo_detection_worker import YoloDetectionWorker


_USE_CURRENT_RESULT = object()


class DetectMain(QWidget):
    detection_requested = Signal(str, str)
    regression_requested = Signal(str, str)
    clear_active_formulas_requested = Signal()
    install_saved_formulas_requested = Signal(object)
    restore_active_formulas_requested = Signal(object)

    def __init__(self):
        super().__init__()
        self.ui = ui_detectmain.Ui_Form()
        self.ui.setupUi(self)
        self._install_resizable_splitters()

        self._calibration_source_path = None
        self._calibration_source_image = None
        self._regression_result = None
        self._regression_dirty = False
        self._detection_result = None
        self._detection_dirty = False
        self._last_completed_result_type = None
        self._last_calibration_directory = None
        self._active_worker_task = None

#        #put the cameraWidget in cameraMainlVLayout
        self.mainCamera = Camera()
#        self.ui.cameraMainGroupBox.setWidget(self.mainCamera._ui.cameraWidget)
        self.ui.cameraMainlVLayout.addWidget(self.mainCamera._ui.cameraWidget)
        self.ui.cameraMainlVLayout.setContentsMargins(0,0,0,0)


        #
        self.layout = QVBoxLayout()
        self.ui.widgetChart.setLayout(self.layout)

        # canvas
#        self.figure = plt.figure()
        self.figure, self.ax = plt.subplots(figsize=(15,15), dpi=100)#nrows=1, ncols=1,figsize=(15,10))#figsize=(20,15))
        self.canvas = FigureCanvasQTAgg(self.figure)
        self._regression_suptitle = None
        self._regression_plot_has_result = False
        self.toolbar = NavigationToolbar2QT(self.canvas, self)
        self.layout.addWidget(self.canvas)
#        self.layout.addWidget(self.toolbar)




        # Linear regression: read source data (full precision for the fit)
        lin_headers, lin_rows = self._read_excel('./interface/linear/table/linear_regression_table.xlsx')
        ycol = self._first_color_index(lin_headers)
        con_lin = [r[lin_headers.index("Con.")] for r in lin_rows]
        y_channel = [r[ycol] for r in lin_rows]
        y_label = lin_headers[ycol]
        channel_color = self.CHANNEL_COLORS.get(y_label, '#2ca02c')
        detect_color = self.CONTRAST_COLORS.get(y_label, '#ff7f0e')

        # Linear regression display table: read from only_table/
        disp_headers, disp_rows = self._read_excel(
            './interface/linear/only_table/linear_regression_table.xlsx')
        self._populate_tableview(self.ui.tabviewOrig, disp_headers, disp_rows)



        # plot the Chart
        font = {'family': 'serif',
                'color':  'black',
                'weight': 'normal',
                'size': 13,
                }

        x = con_lin
        y = y_channel
        slope, intercept, r, p, std_err = stats.linregress(x, y)
        R2 = pow(r, 2)
        def myfunc(x):
            return slope*x+intercept
        mymodel = list(map(myfunc, x))

        text = "Y = {:.4f}*X+{:.2f}".format(slope, intercept)
#        print('text=', text)
#        print("y = {:.2f}*x+{:.2f}".format(slope, intercept))

        # Detection: read detection-source data, compute predicted concentrations
        det_headers, det_rows = self._read_excel('./interface/detect/table/detection_table.xlsx')
        dcol = self._first_color_index(det_headers)
        det_channel = [r[dcol] for r in det_rows]
        con_pred = [(v - intercept) / slope for v in det_channel]
        print('con: ', con_pred)

#        plt.plot(x, y, 'k')
        plt.title('Linear Regression and Detection of Real Samples', fontdict=font, fontsize=15) # 'Linear Regression and ML-assisted HT-Detection'

        # Formula overlay: keep clear of the regression line by docking to the
        # corner opposite the slope direction; axes-fraction coords keep it
        # stable across data scales.
        if slope >= 0:
            tx, ha = 0.02, 'left'
        else:
            tx, ha = 0.98, 'right'
        self.ax.text(tx, 0.96,
                "Y = {:.4f} * X+{:.2f} ,  R$^2$ = {:.4f}".format(slope, intercept, R2),
                transform=self.ax.transAxes,
                ha=ha, va='top',
                fontsize=16, fontstyle='italic', fontfamily='times new roman',
                color=(0, 0, 0, 1),
                bbox=dict(facecolor=channel_color, alpha=0.5, edgecolor='none',
                          boxstyle='round,pad=0.3'))
        # plt.text(61, 143, r'$\cos(2 \pi t) \exp(-t)$', fontdict=font)
        plt.xlabel('Concentration of AA (μM)', fontdict=font, fontsize=13) #'Concentration of Hg$^{2+}$ (μM)'
        plt.ylabel('{} Value'.format(y_label), fontdict=font, fontsize=13)



        plt.scatter(x, y, color=channel_color, linewidths=4, zorder=1)
        plt.plot(x, mymodel, color=channel_color, linewidth=3, linestyle='--', alpha=0.7, zorder=2)
        plt.scatter(con_pred, det_channel, color=detect_color, linewidths=4, zorder=3)




        plt.legend(('experimental data', 'linear regression', 'detection result'),
                   loc='lower right', shadow=True)

        marker_color = detect_color
        marker_size = 9   # one size smaller than before
        # 28 px between bands comfortably exceeds the rendered label-box height
        # (~19 px at fontsize 9 with pad=1) so adjacent-band labels cannot overlap
        # vertically. Anchored at each point's x preserves left-to-right order; x
        # collisions bump the label to the next band. label_half_w stays pessimistic
        # so tight clusters still fan out across distinct bands.
        dx_global = -3
        order = sorted(range(len(con_pred)), key=lambda i: con_pred[i])
        x_span = (max(con_pred) - min(con_pred)) if len(con_pred) > 1 else 1.0
        label_half_w = max(x_span * 0.11, 1e-6)
        dy_cycle = [-26, -54, -82, -110, 26, 54, 82, 110]
        band_intervals = {dy: [] for dy in dy_cycle}
        assigned_dy = [dy_cycle[0]] * len(con_pred)
        for idx in order:
            xL = con_pred[idx] - label_half_w
            xR = con_pred[idx] + label_half_w
            for dy in dy_cycle:
                collides = any(max(xL, l) < min(xR, r) for l, r in band_intervals[dy])
                if not collides:
                    band_intervals[dy].append((xL, xR))
                    assigned_dy[idx] = dy
                    break
        for i, (cx, cy) in enumerate(zip(con_pred, det_channel)):
            plt.annotate('({:.2f},{:.2f})'.format(cx, cy),
                         xy=(cx, cy), xytext=(dx_global, assigned_dy[i]),
                         textcoords='offset points',
                         ha='center', va='center',
                         fontsize=marker_size, color=marker_color,
                         bbox=dict(facecolor='white', alpha=0.75,
                                   edgecolor='none', pad=1.0),
                         arrowprops=dict(arrowstyle='-', color='lightgray',
                                         lw=0.8, linestyle='--', alpha=0.9))

        # Modest axis padding so labels stay inside the plot frame at any window size.
        y0, y1 = self.ax.get_ylim()
        yr = y1 - y0
        self.ax.set_ylim(y0 - yr * 0.12, y1 + yr * 0.06)


        # Render into the embedded Qt canvas only — no standalone pyplot window.
        self.canvas.draw()



        # Detection display table: read from only_table/
        det_disp_headers, det_disp_rows = self._read_excel(
            './interface/detect/only_table/detection_table.xlsx')
        self._populate_tableview(self.ui.tabviewRecg, det_disp_headers, det_disp_rows)
        self.ui.tabviewRecg.horizontalHeader().setSectionResizeMode(
            QHeaderView.Stretch
        )


        self.origImg = self._resolve_image('./interface/linear/image/linear_regression_image')
        self.recgImg = self._resolve_image('./interface/detect/image/detection_image')
        self._origPixmap = QPixmap(self.origImg)
        self._recgPixmap = QPixmap(self.recgImg)
        for lbl in (self.ui.labelOrigImg, self.ui.labelRecgImg):
            lbl.setMinimumSize(1, 1)
            lbl.setAlignment(Qt.AlignCenter)
            lbl.setScaledContents(False)
            # Event filter catches the label's own resize, which fires reliably
            # even when DetectMain itself isn't in a layout (the UI widget is).
            lbl.installEventFilter(self)
        self._update_image_scales()

        self.ui.progressBar.setValue(100)
        try:
            with open('./interface/detect/time.txt', 'r', encoding='utf-8') as f:
                raw = f.read().strip()
            digits = ''.join(ch for ch in raw if ch.isdigit())
            if digits:
                self.ui.lcdNumber.display(int(digits))
        except OSError:
            pass

        self._detection_thread = QThread(self)
        self._detection_worker = YoloDetectionWorker()
        self._detection_worker.moveToThread(self._detection_thread)
        self.detection_requested.connect(self._detection_worker.detect)
        self.regression_requested.connect(self._detection_worker.regress)
        self.clear_active_formulas_requested.connect(
            self._detection_worker.clear_active_formulas
        )
        self.install_saved_formulas_requested.connect(
            self._detection_worker.install_saved_formulas
        )
        self.restore_active_formulas_requested.connect(
            self._detection_worker.restore_active_formulas
        )
        self._detection_worker.finished.connect(self._on_detection_finished)
        self._detection_worker.failed.connect(self._on_detection_failed)
        self._detection_worker.regression_finished.connect(
            self._on_regression_finished
        )
        self._detection_worker.regression_failed.connect(self._on_regression_failed)
        self._detection_thread.finished.connect(self._detection_worker.deleteLater)
        self._detection_thread.start()

        self.ui.pushButton_5.clicked.connect(self._select_calibration_image)
        self.ui.pushButton_4.clicked.connect(self._start_linear_regression)
        self.ui.pushButton.clicked.connect(self._select_detection_image)
        self.ui.pushButton_7.clicked.connect(self._plot_regression_result)
        self.ui.pushButton_8.clicked.connect(self._save_pending_result)
        self._reset_calibration_ui()
        app = QCoreApplication.instance()
        if app is not None:
            app.aboutToQuit.connect(self._shutdown_detection_thread)

    CHANNEL_COLORS = {
        'Red':   '#d62728',
        'Green': '#2ca02c',
        'Blue':  '#1f77b4',
    }

    # High-contrast partners for each regression channel: complementary-ish hues
    # that stay clearly separated from the corresponding CHANNEL_COLORS entry.
    CONTRAST_COLORS = {
        'Red':   '#17becf',   # cyan vs red
        'Green': '#e377c2',   # magenta vs green
        'Blue':  '#ff7f0e',   # orange vs blue
    }

    IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png')

    CALIBRATION_TABLE_HEADERS = ("No.", "Con.", "Red", "Green", "Blue")
    CALIBRATION_PLOT_PLACEHOLDER = "Run linear regression, then click Plot."
    REGRESSION_CHANNEL_FIELDS = {"R": "Red", "G": "Green", "B": "Blue"}
    FORMULA_CHANNELS = ("R", "G", "B")
    FORMULA_FIELDS = ("slope", "intercept", "r", "R2", "p", "std_err")
    LINEAR_EXPORT_FILENAMES = (
        "linear_con_rgb.xlsx",
        "standard_curve.png",
        "calibration_annotated.png",
    )
    DETECTION_TABLE_HEADERS = (
        "Name", "No.", "Con.", "Red", "Green", "Blue", None,
        "x0_con", "y0_con", "x1_con", "y1_con", "w_con", "h_con",
    )
    MAX_DETECTION_STEM_LENGTH = 100

    @staticmethod
    def _decode_calibration_image(image_path):
        import cv2
        import numpy as np

        path = Path(image_path)
        if not path.is_file():
            raise ValueError("The selected calibration image does not exist or is not a file.")

        try:
            encoded = np.fromfile(str(path), dtype=np.uint8)
        except OSError as error:
            raise ValueError("The selected calibration image could not be read: {}".format(error)) from error
        if encoded.size == 0:
            raise ValueError("The selected calibration image is empty.")

        image = cv2.imdecode(encoded, cv2.IMREAD_COLOR)
        if image is None:
            raise ValueError("The selected file is not a valid supported calibration image.")
        if image.ndim != 3 or image.shape[0] <= 0 or image.shape[1] <= 0:
            raise ValueError("The decoded calibration image has invalid dimensions.")
        return image

    @staticmethod
    def _bgr_image_to_pixmap(image):
        height, width, channels = image.shape
        bytes_per_line = channels * width
        qt_image = QImage(
            image.data, width, height, bytes_per_line, QImage.Format_RGB888
        ).rgbSwapped().copy()
        return QPixmap.fromImage(qt_image)

    def _show_calibration_plot_placeholder(self):
        if self._regression_suptitle is not None:
            self._regression_suptitle.remove()
            self._regression_suptitle = None
        self.ax.clear()
        self.ax.text(
            0.5,
            0.5,
            self.CALIBRATION_PLOT_PLACEHOLDER,
            transform=self.ax.transAxes,
            ha="center",
            va="center",
        )
        self.ax.set_axis_off()
        self.canvas.draw()
        self._regression_plot_has_result = False

    def _clear_calibration_results(self):
        self._regression_result = None
        self._regression_dirty = False
        self._populate_tableview(
            self.ui.tabviewOrig, self.CALIBRATION_TABLE_HEADERS, []
        )
        self._show_calibration_plot_placeholder()
        self.ui.pushButton_7.setEnabled(False)
        self._update_save_button()

    def _validated_linear_export_payload(self, payload=_USE_CURRENT_RESULT):
        if payload is _USE_CURRENT_RESULT:
            payload = self._regression_result
        if not isinstance(payload, dict):
            raise ValueError("No linear regression result is available.")
        source_path = payload.get("source_path")
        if not isinstance(source_path, str) or not source_path.strip():
            raise ValueError("The regression source path is missing.")
        if not Path(source_path).is_file():
            raise ValueError("The regression source image no longer exists: {}".format(source_path))

        image = payload.get("image")
        if not isinstance(image, np.ndarray) or image.ndim != 3 or image.size == 0:
            raise ValueError("The regression annotated image is invalid.")
        if image.shape[2] not in (3, 4):
            raise ValueError("The regression annotated image has an unsupported channel count.")

        samples = payload.get("samples")
        if not isinstance(samples, list) or not samples:
            raise ValueError("The regression samples are missing or invalid.")
        validated_samples = []
        included_count = 0
        included_concentrations = set()
        for index, sample in enumerate(samples, start=1):
            if not isinstance(sample, dict):
                raise ValueError("Regression sample {} is invalid.".format(index))
            row = {}
            for field in self.CALIBRATION_TABLE_HEADERS:
                value = sample.get(field)
                if isinstance(value, bool) or not isinstance(value, numbers.Real):
                    raise ValueError("Regression sample {} field {} must be numeric.".format(index, field))
                value = float(value)
                if not math.isfinite(value):
                    raise ValueError("Regression sample {} field {} must be finite.".format(index, field))
                row[field] = sample[field]
            included = sample.get("included")
            if not isinstance(included, bool):
                raise ValueError("Regression sample {} included must be bool.".format(index))
            row["included"] = included
            included_count += int(included)
            if included:
                included_concentrations.add(float(row["Con."]))
            validated_samples.append(row)
        if included_count < 2:
            raise ValueError("At least two regression samples must be included.")
        if len(included_concentrations) < 2:
            raise ValueError("Included samples must contain at least two distinct concentrations.")

        formulas = payload.get("formulas")
        if not isinstance(formulas, dict) or set(formulas) != set(self.FORMULA_CHANNELS):
            raise ValueError("Regression formulas must contain exactly R, G, and B.")
        validated_formulas = {}
        for channel in self.FORMULA_CHANNELS:
            formula = formulas[channel]
            if not isinstance(formula, dict):
                raise ValueError("Regression formula {} is invalid.".format(channel))
            values = {}
            for field in self.FORMULA_FIELDS:
                value = formula.get(field)
                if isinstance(value, bool) or not isinstance(value, numbers.Real):
                    raise ValueError("Regression formula {}.{} must be numeric.".format(channel, field))
                value = float(value)
                if not math.isfinite(value):
                    raise ValueError("Regression formula {}.{} must be finite.".format(channel, field))
                values[field] = formula[field]
            if math.isclose(float(values["slope"]), 0.0, rel_tol=0.0, abs_tol=1e-12):
                raise ValueError("Regression formula {} slope must not be zero or near zero.".format(channel))
            validated_formulas[channel] = values

        selected_channel = payload.get("selected_channel")
        if selected_channel not in self.FORMULA_CHANNELS:
            raise ValueError("The selected regression channel is invalid.")
        return {
            "source_path": source_path,
            "image": image,
            "samples": validated_samples,
            "formulas": validated_formulas,
            "selected_channel": selected_channel,
        }

    def _has_valid_linear_export(self):
        try:
            self._validated_linear_export_payload()
        except ValueError:
            return False
        return True

    def _pending_save_type(self):
        if self._last_completed_result_type == "linear" and self._regression_dirty:
            return "linear"
        if self._last_completed_result_type == "detection" and self._detection_dirty:
            return "detection"
        if self._regression_dirty:
            return "linear"
        if self._detection_dirty:
            return "detection"
        return None

    def _update_save_button(self):
        pending = self._pending_save_type()
        if pending == "linear":
            text = "Save Linear"
            available = self._has_valid_linear_export()
        elif pending == "detection":
            text = "Save Detection"
            available = self._has_valid_detection_export()
        else:
            text = "Save"
            available = False
        self.ui.pushButton_8.setText(text)
        self.ui.pushButton_8.setEnabled(available and self._active_worker_task is None)

    @Slot()
    def _save_pending_result(self):
        pending = self._pending_save_type()
        if pending == "linear":
            self._save_linear_result()
        elif pending == "detection":
            self._save_detection_result()

    def _confirm_discard_unsaved_regression(self):
        if not self._regression_dirty:
            return True
        answer = QMessageBox.question(
            self,
            "Unsaved linear regression",
            "当前线性回归结果尚未保存，继续操作将丢弃该结果。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        return answer == QMessageBox.Yes

    @classmethod
    def _safe_detection_stem(cls, source_path):
        stem = Path(source_path).stem
        invalid = set('<>:"/\\|?*')
        sanitized = "".join(
            "_" if character in invalid or ord(character) < 32 or ord(character) == 127 else character
            for character in stem
        ).rstrip(" .")
        if not sanitized:
            sanitized = "detection_result"
        reserved = {"CON", "PRN", "AUX", "NUL"}
        reserved.update("COM{}".format(index) for index in range(1, 10))
        reserved.update("LPT{}".format(index) for index in range(1, 10))
        if sanitized.split(".", 1)[0].upper() in reserved:
            sanitized = "_" + sanitized
        sanitized = sanitized[:cls.MAX_DETECTION_STEM_LENGTH].rstrip(" .")
        return sanitized or "detection_result"

    def _validated_detection_export_payload(self, payload=_USE_CURRENT_RESULT):
        if payload is _USE_CURRENT_RESULT:
            payload = self._detection_result
        if not isinstance(payload, dict):
            raise ValueError("No detection result is available.")
        source_path = payload.get("source_path")
        if not isinstance(source_path, str) or not source_path.strip():
            raise ValueError("The detection source path is missing.")
        if not Path(source_path).is_file():
            raise ValueError("The detection source image no longer exists: {}".format(source_path))
        image = payload.get("image")
        if not isinstance(image, np.ndarray) or image.ndim != 3 or image.size == 0:
            raise ValueError("The detection annotated image is invalid.")
        if image.shape[2] not in (3, 4):
            raise ValueError("The detection annotated image has an unsupported channel count.")
        if image.shape[2] != 3:
            raise ValueError("The detection annotated image must be a three-channel BGR image.")
        targets = payload.get("targets")
        if not isinstance(targets, list) or not targets:
            raise ValueError("At least one valid detection target is required.")

        validated_targets = []
        numeric_fields = ("No.", "Con.", "Red", "Green", "Blue")
        for index, target in enumerate(targets, start=1):
            if not isinstance(target, dict):
                raise ValueError("Detection target {} is invalid.".format(index))
            validated = {}
            for field in numeric_fields:
                value = target.get(field)
                if isinstance(value, bool) or not isinstance(value, numbers.Real):
                    raise ValueError("Detection target {} field {} must be numeric.".format(index, field))
                if not math.isfinite(float(value)):
                    raise ValueError("Detection target {} field {} must be finite.".format(index, field))
                validated[field] = value
            roi = target.get("rgb_roi")
            if not isinstance(roi, (tuple, list)) or len(roi) != 4:
                raise ValueError("Detection target {} rgb_roi is invalid.".format(index))
            coordinates = []
            for value in roi:
                if isinstance(value, bool) or not isinstance(value, numbers.Real):
                    raise ValueError("Detection target {} rgb_roi must be numeric.".format(index))
                if not math.isfinite(float(value)):
                    raise ValueError("Detection target {} rgb_roi must be finite.".format(index))
                coordinates.append(value)
            x0, y0, x1, y1 = coordinates
            if x1 <= x0 or y1 <= y0:
                raise ValueError("Detection target {} rgb_roi must have positive dimensions.".format(index))
            validated["rgb_roi"] = tuple(coordinates)
            validated_targets.append(validated)
        return {
            "source_path": source_path,
            "source_stem": Path(source_path).stem,
            "safe_stem": self._safe_detection_stem(source_path),
            "image": image,
            "targets": validated_targets,
        }

    def _has_valid_detection_export(self):
        try:
            self._validated_detection_export_payload()
        except ValueError:
            return False
        return True

    def _is_valid_detection_payload(self, payload):
        previous = self._detection_result
        self._detection_result = payload
        try:
            return self._has_valid_detection_export()
        finally:
            self._detection_result = previous

    @classmethod
    def _build_detection_workbook_bytes(cls, export_payload):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        for column, header in enumerate(cls.DETECTION_TABLE_HEADERS, start=1):
            if header is not None:
                worksheet.cell(1, column, header)
        for row_index, target in enumerate(export_payload["targets"], start=2):
            x0, y0, x1, y1 = target["rgb_roi"]
            values = (
                export_payload["source_stem"], target["No."], target["Con."],
                target["Red"], target["Green"], target["Blue"], None,
                x0, y0, x1, y1, x1 - x0, y1 - y0,
            )
            for column, value in enumerate(values, start=1):
                if value is not None:
                    worksheet.cell(row_index, column, value)
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    @classmethod
    def _validate_detection_workbook_bytes(cls, workbook_bytes, export_payload):
        def values_match(actual, expected):
            if isinstance(expected, numbers.Real) and not isinstance(expected, bool):
                return (
                    isinstance(actual, numbers.Real)
                    and not isinstance(actual, bool)
                    and math.isclose(float(actual), float(expected), rel_tol=1e-15, abs_tol=0.0)
                )
            return actual == expected

        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        try:
            if workbook.sheetnames != ["Sheet1"]:
                raise ValueError("The detection workbook must contain only Sheet1.")
            worksheet = workbook["Sheet1"]
            headers = tuple(worksheet.cell(1, column).value for column in range(1, 14))
            if headers != cls.DETECTION_TABLE_HEADERS:
                raise ValueError("The detection workbook headers A1:M1 are invalid.")
            for row_index, target in enumerate(export_payload["targets"], start=2):
                x0, y0, x1, y1 = target["rgb_roi"]
                expected = (
                    export_payload["source_stem"], target["No."], target["Con."],
                    target["Red"], target["Green"], target["Blue"], None,
                    x0, y0, x1, y1, x1 - x0, y1 - y0,
                )
                actual = tuple(worksheet.cell(row_index, column).value for column in range(1, 14))
                if not all(values_match(value, wanted) for value, wanted in zip(actual, expected)):
                    raise ValueError("Detection workbook row {} does not match the payload.".format(row_index))
                if worksheet.cell(row_index, 7).value is not None:
                    raise ValueError("Detection workbook column G must remain empty.")
        finally:
            workbook.close()

    @classmethod
    def _build_detection_export_bytes(cls, export_payload):
        workbook_bytes = cls._build_detection_workbook_bytes(export_payload)
        cls._validate_detection_workbook_bytes(workbook_bytes, export_payload)
        import cv2
        success, encoded = cv2.imencode(".png", export_payload["image"])
        if not success or encoded.size == 0:
            raise ValueError("The detection annotated PNG could not be encoded.")
        png_bytes = encoded.tobytes()
        decoded = cv2.imdecode(np.frombuffer(png_bytes, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
        if decoded is None or not np.array_equal(decoded, export_payload["image"]):
            raise ValueError("The detection annotated PNG failed color verification.")
        return workbook_bytes, png_bytes

    @staticmethod
    def _commit_detection_export(target_directory, safe_stem, workbook_bytes, png_bytes):
        target_directory.mkdir(parents=True, exist_ok=True)
        transaction_path = Path(tempfile.mkdtemp(prefix=".save-detection-", dir=str(target_directory)))
        staged_xlsx = transaction_path / "result.xlsx"
        staged_png = transaction_path / "result.png"
        preserve_transaction = False
        committed_pair = None
        try:
            staged_xlsx.write_bytes(workbook_bytes)
            staged_png.write_bytes(png_bytes)
            if staged_xlsx.read_bytes() != workbook_bytes or staged_png.read_bytes() != png_bytes:
                raise RuntimeError("Detection staging verification failed.")
            index = 0
            while True:
                candidate = safe_stem if index == 0 else "{}_{}".format(safe_stem, index)
                final_xlsx = target_directory / "{}.xlsx".format(candidate)
                final_png = target_directory / "{}.png".format(candidate)
                if final_xlsx.exists() or final_png.exists():
                    index += 1
                    continue
                try:
                    os.link(str(staged_xlsx), str(final_xlsx))
                except FileExistsError:
                    index += 1
                    continue
                try:
                    os.link(str(staged_png), str(final_png))
                except FileExistsError:
                    try:
                        final_xlsx.unlink()
                    except OSError as rollback_error:
                        preserve_transaction = True
                        raise RuntimeError(
                            "A competing file was detected and rollback failed. Recovery files are in {}: {}".format(
                                transaction_path, rollback_error
                            )
                        ) from rollback_error
                    index += 1
                    continue
                except Exception as error:
                    try:
                        final_xlsx.unlink()
                    except OSError as rollback_error:
                        preserve_transaction = True
                        raise RuntimeError(
                            "Detection save failed and rollback was incomplete. Recovery files are in {}: {}".format(
                                transaction_path, rollback_error
                            )
                        ) from error
                    raise RuntimeError("Detection PNG commit failed; the Excel file was removed: {}".format(error)) from error
                committed_pair = (final_xlsx, final_png)
                break
        except Exception as error:
            if transaction_path.exists() and not preserve_transaction:
                try:
                    shutil.rmtree(transaction_path)
                except OSError as cleanup_error:
                    raise RuntimeError(
                        "Detection save failed and transaction cleanup was incomplete. Recovery files are in {}: {}".format(
                            transaction_path, cleanup_error
                        )
                    ) from error
            raise
        try:
            shutil.rmtree(transaction_path)
        except OSError as cleanup_error:
            raise RuntimeError(
                "Detection files were committed, but transaction cleanup failed. Recovery files are in {}: {}".format(
                    transaction_path, cleanup_error
                )
            ) from cleanup_error
        return committed_pair

    @Slot()
    def _save_detection_result(self):
        if self._active_worker_task is not None:
            return
        try:
            export_payload = self._validated_detection_export_payload()
        except ValueError as error:
            self._detection_dirty = bool(self._detection_result)
            self._update_save_button()
            QMessageBox.critical(self, "Save Detection error", str(error))
            return
        repository_root = Path(__file__).resolve().parent.parent
        target_directory = repository_root / "HT-Detector_Peng" / "runs" / "detect" / "results" / "detection"
        self._set_active_worker_task("save_detection")
        try:
            workbook_bytes, png_bytes = self._build_detection_export_bytes(export_payload)
            final_xlsx, final_png = self._commit_detection_export(
                target_directory, export_payload["safe_stem"], workbook_bytes, png_bytes
            )
        except Exception as error:
            self._detection_dirty = True
            self._set_active_worker_task(None)
            QMessageBox.critical(self, "Save Detection error", str(error))
            return
        self._detection_dirty = False
        self._set_active_worker_task(None)
        QMessageBox.information(
            self,
            "Save Detection",
            "Detection results were saved to:\n{}\n{}".format(final_xlsx, final_png),
        )

    @classmethod
    def _build_linear_workbook_bytes(cls, export_payload):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        sample_headers = ("No.", "Con.", "Red", "Green", "Blue", "included")
        for column, header in enumerate(sample_headers, start=1):
            worksheet.cell(1, column, header)
        for row_index, sample in enumerate(export_payload["samples"], start=2):
            for column, header in enumerate(sample_headers, start=1):
                worksheet.cell(row_index, column, sample[header])

        formula_headers = ("Channel",) + cls.FORMULA_FIELDS
        for column, header in enumerate(formula_headers, start=8):
            worksheet.cell(1, column, header)
        for row_index, channel in enumerate(cls.FORMULA_CHANNELS, start=2):
            worksheet.cell(row_index, 8, channel)
            for column, field in enumerate(cls.FORMULA_FIELDS, start=9):
                worksheet.cell(row_index, column, export_payload["formulas"][channel][field])

        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        return buffer.getvalue()

    @classmethod
    def _validate_linear_workbook_bytes(cls, workbook_bytes, export_payload):
        def values_match(actual, expected):
            if isinstance(expected, bool):
                return isinstance(actual, bool) and actual is expected
            if isinstance(expected, numbers.Real) and not isinstance(expected, bool):
                return (
                    isinstance(actual, numbers.Real)
                    and not isinstance(actual, bool)
                    and math.isclose(
                        float(actual), float(expected), rel_tol=1e-15, abs_tol=0.0
                    )
                )
            return actual == expected

        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
        try:
            if workbook.sheetnames != ["Sheet1"]:
                raise ValueError("The exported workbook must contain only Sheet1.")
            worksheet = workbook["Sheet1"]
            sample_headers = tuple(worksheet.cell(1, column).value for column in range(1, 7))
            if sample_headers != ("No.", "Con.", "Red", "Green", "Blue", "included"):
                raise ValueError("The exported sample headers A1:F1 are invalid.")
            if worksheet["G1"].value is not None:
                raise ValueError("Column G must remain empty.")
            formula_headers = tuple(worksheet.cell(1, column).value for column in range(8, 15))
            if formula_headers != ("Channel",) + cls.FORMULA_FIELDS:
                raise ValueError("The exported formula headers H1:N1 are invalid.")
            for row_index in range(1, max(len(export_payload["samples"]) + 1, 4) + 1):
                if worksheet.cell(row_index, 7).value is not None:
                    raise ValueError("Column G must remain empty.")
            for row_index, expected in enumerate(export_payload["samples"], start=2):
                actual = tuple(worksheet.cell(row_index, column).value for column in range(1, 7))
                wanted = tuple(expected[field] for field in ("No.", "Con.", "Red", "Green", "Blue", "included"))
                if not all(values_match(value, expected) for value, expected in zip(actual, wanted)):
                    raise ValueError("The exported sample row {} does not match the payload.".format(row_index))
                if worksheet.cell(row_index, 7).value is not None:
                    raise ValueError("Column G must remain empty.")
            for row_index, channel in enumerate(cls.FORMULA_CHANNELS, start=2):
                actual = tuple(worksheet.cell(row_index, column).value for column in range(8, 15))
                wanted = (channel,) + tuple(export_payload["formulas"][channel][field] for field in cls.FORMULA_FIELDS)
                if not all(values_match(value, expected) for value, expected in zip(actual, wanted)):
                    raise ValueError("The exported {} formula does not match the payload.".format(channel))
        finally:
            workbook.close()

    def _build_linear_export_bytes(self, export_payload):
        self._plot_regression_result()
        workbook_bytes = self._build_linear_workbook_bytes(export_payload)
        self._validate_linear_workbook_bytes(workbook_bytes, export_payload)

        curve_buffer = io.BytesIO()
        self.canvas.print_png(curve_buffer)
        curve_bytes = curve_buffer.getvalue()
        if not curve_bytes.startswith(b"\x89PNG\r\n\x1a\n"):
            raise ValueError("The standard curve PNG could not be encoded.")

        import cv2
        decoded_curve = cv2.imdecode(np.frombuffer(curve_bytes, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
        if decoded_curve is None or decoded_curve.size == 0:
            raise ValueError("The standard curve PNG failed verification.")
        success, encoded = cv2.imencode(".png", export_payload["image"])
        if not success or encoded.size == 0:
            raise ValueError("The calibration annotated PNG could not be encoded.")
        annotated_bytes = encoded.tobytes()
        decoded = cv2.imdecode(np.frombuffer(annotated_bytes, dtype=np.uint8), cv2.IMREAD_UNCHANGED)
        if decoded is None or decoded.shape != export_payload["image"].shape:
            raise ValueError("The calibration annotated PNG failed verification.")
        return {
            "linear_con_rgb.xlsx": workbook_bytes,
            "standard_curve.png": curve_bytes,
            "calibration_annotated.png": annotated_bytes,
        }

    @staticmethod
    def _commit_linear_export(target_directory, files):
        target_directory.mkdir(parents=True, exist_ok=True)
        transaction_path = Path(tempfile.mkdtemp(prefix=".save-linear-", dir=str(target_directory)))
        staged_path = transaction_path / "staged"
        backup_path = transaction_path / "backup"
        staged_path.mkdir()
        backup_path.mkdir()
        backups = []
        committed = []
        try:
            for filename, content in files.items():
                (staged_path / filename).write_bytes(content)
            for filename in files:
                target = target_directory / filename
                if target.exists():
                    os.replace(str(target), str(backup_path / filename))
                    backups.append(filename)
            for filename in files:
                os.replace(str(staged_path / filename), str(target_directory / filename))
                committed.append(filename)
        except Exception as error:
            rollback_errors = []
            for filename in reversed(committed):
                try:
                    (target_directory / filename).unlink()
                except OSError as rollback_error:
                    rollback_errors.append("remove {}: {}".format(filename, rollback_error))
            for filename in reversed(backups):
                try:
                    os.replace(str(backup_path / filename), str(target_directory / filename))
                except OSError as rollback_error:
                    rollback_errors.append("restore {}: {}".format(filename, rollback_error))
            if rollback_errors:
                raise RuntimeError(
                    "Save failed and rollback was incomplete. Recovery files are in {}. {}".format(
                        transaction_path, "; ".join(rollback_errors)
                    )
                ) from error
            shutil.rmtree(transaction_path, ignore_errors=True)
            raise RuntimeError("Save failed; all previous files were restored: {}".format(error)) from error
        shutil.rmtree(transaction_path)

    @Slot()
    def _save_linear_result(self):
        if self._active_worker_task is not None:
            return
        try:
            export_payload = self._validated_linear_export_payload()
        except ValueError as error:
            self._regression_dirty = bool(self._regression_result)
            self._update_save_button()
            QMessageBox.critical(self, "Save Linear error", str(error))
            return

        repository_root = Path(__file__).resolve().parent.parent
        target_directory = repository_root / "HT-Detector_Peng" / "runs" / "detect" / "results" / "linear"
        existing = [name for name in self.LINEAR_EXPORT_FILENAMES if (target_directory / name).exists()]
        if existing:
            answer = QMessageBox.question(
                self,
                "Overwrite linear regression files",
                "The following files in {} will be replaced:\n\n{}".format(
                    target_directory, "\n".join(existing)
                ),
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                self._update_save_button()
                return

        self._set_active_worker_task("save_linear")
        try:
            files = self._build_linear_export_bytes(export_payload)
            self._commit_linear_export(target_directory, files)
        except Exception as error:
            self._regression_dirty = True
            self._set_active_worker_task(None)
            QMessageBox.critical(self, "Save Linear error", str(error))
            return

        self.install_saved_formulas_requested.emit(export_payload["formulas"])
        self._regression_dirty = False
        self._set_active_worker_task(None)
        QMessageBox.information(
            self, "Save Linear", "Linear regression results were saved to:\n{}".format(target_directory)
        )

    def _regression_plot_data(self):
        payload = self._regression_result
        if not isinstance(payload, dict):
            raise ValueError("No linear regression result is available.")

        samples = payload.get("samples")
        formulas = payload.get("formulas")
        channel = payload.get("selected_channel")
        if not isinstance(samples, (list, tuple)) or not samples:
            raise ValueError("The regression result contains no samples.")
        if channel not in self.REGRESSION_CHANNEL_FIELDS:
            raise ValueError("The regression result has an invalid selected channel.")
        if not isinstance(formulas, dict) or not isinstance(formulas.get(channel), dict):
            raise ValueError("The regression formula for channel {} is missing.".format(channel))

        formula = formulas[channel]
        try:
            slope = float(formula["slope"])
            intercept = float(formula["intercept"])
            r_squared = float(formula["R2"])
        except (KeyError, TypeError, ValueError) as error:
            raise ValueError("The regression formula is incomplete or invalid.") from error
        if not all(np.isfinite(value) for value in (slope, intercept, r_squared)):
            raise ValueError("The regression formula contains non-finite values.")

        channel_field = self.REGRESSION_CHANNEL_FIELDS[channel]
        plot_samples = []
        for sample in samples:
            if not isinstance(sample, dict):
                raise ValueError("The regression result contains an invalid sample.")
            try:
                concentration = float(sample["Con."])
                intensity = float(sample[channel_field])
            except (KeyError, TypeError, ValueError) as error:
                raise ValueError(
                    "A regression sample is missing valid concentration or {} data.".format(
                        channel_field
                    )
                ) from error
            if not np.isfinite(concentration) or not np.isfinite(intensity):
                raise ValueError("A regression sample contains non-finite plot data.")
            plot_samples.append((concentration, intensity, sample.get("included") is True))

        included = [sample for sample in plot_samples if sample[2]]
        if len(included) < 2 or len({sample[0] for sample in included}) < 2:
            raise ValueError("At least two distinct included concentrations are required to plot.")
        return channel, channel_field, plot_samples, included, slope, intercept, r_squared

    def _has_valid_regression_result(self):
        try:
            self._regression_plot_data()
        except ValueError:
            return False
        return True

    @Slot()
    def _plot_regression_result(self):
        try:
            (channel, channel_field, _plot_samples, included,
             slope, intercept, r_squared) = self._regression_plot_data()
        except ValueError as error:
            self._show_calibration_plot_placeholder()
            self.ui.pushButton_7.setEnabled(False)
            self._update_save_button()
            QMessageBox.critical(self, "Plot error", str(error))
            return

        used_x = [sample[0] for sample in included]
        used_y = [sample[1] for sample in included]
        color = self.CHANNEL_COLORS[channel_field]

        self.ax.clear()
        self.ax.set_axis_on()
        self.ax.scatter(
            used_x, used_y, color=color, marker="o", s=55,
            label="Used in regression (n={})".format(len(included)), zorder=3,
        )
        line_x = np.array([min(used_x), max(used_x)], dtype=float)
        line_y = slope * line_x + intercept
        self.ax.plot(
            line_x, line_y, color=color,
            linewidth=2.2, label="Linear fit", zorder=2,
        )
        sign = "+" if intercept >= 0 else "-"
        equation = "y = {:.4f}x {} {:.4f}".format(slope, sign, abs(intercept))
        title = "Linear Regression – {} Channel\n{}    R² = {:.4f}".format(
            channel_field, equation, r_squared
        )
        if self._regression_suptitle is None:
            self._regression_suptitle = self.figure.suptitle(title, y=0.97)
        else:
            self._regression_suptitle.set_text(title)
        self.figure.subplots_adjust(left=0.14, right=0.96, bottom=0.16, top=0.78)
        self.ax.set_xlabel("Concentration")
        self.ax.set_ylabel("{} Intensity".format(channel_field))
        self.ax.grid(True, alpha=0.3)
        self.ax.legend()

        def padded_limits(values):
            lower = float(min(values))
            upper = float(max(values))
            span = upper - lower
            margin = span * 0.05 if span > 0 else max(abs(lower) * 0.05, 0.5)
            return lower - margin, upper + margin

        # Limits are rebuilt from the current included points and fit endpoints
        # after every clear; excluded samples and previous axes state cannot affect them.
        self.ax.set_xlim(padded_limits(list(used_x) + list(line_x)))
        self.ax.set_ylim(padded_limits(list(used_y) + list(line_y)))
        self.canvas.draw()
        self._regression_plot_has_result = True
        self.ui.pushButton_7.setEnabled(True)
        self._update_save_button()

    def _set_active_worker_task(self, task):
        if task not in (None, "detection", "regression", "save_linear", "save_detection"):
            raise ValueError("Unknown worker task: {}".format(task))
        self._active_worker_task = task
        busy = task is not None
        self.ui.pushButton_5.setEnabled(not busy)
        self.ui.pushButton_4.setEnabled(
            not busy and self._calibration_source_path is not None
        )
        self.ui.pushButton.setEnabled(not busy)
        self.ui.pushButton_7.setEnabled(
            not busy and self._has_valid_regression_result()
        )
        self._update_save_button()

    def _reset_calibration_ui(self):
        self._calibration_source_path = None
        self._calibration_source_image = None
        self.origImg = None
        self._origPixmap = QPixmap()
        self.ui.labelOrigImg.clear()
        self.ui.labelOrigImg.setText("Import a calibration image")
        self.ui.labelOrigImg.setAlignment(Qt.AlignCenter)
        self._clear_calibration_results()
        self._set_active_worker_task(None)

    @Slot()
    def _select_calibration_image(self):
        initial_directory = self._last_calibration_directory or ""
        image_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select a calibration image",
            initial_directory,
            "Images (*.jpg *.jpeg *.png *.bmp *.tif *.tiff)",
        )
        if not image_path:
            return

        try:
            source_image = self._decode_calibration_image(image_path)
            normalized_path = str(Path(image_path).resolve(strict=True))
            pixmap = self._bgr_image_to_pixmap(source_image)
            if pixmap.isNull():
                raise ValueError("The decoded calibration image could not be displayed.")
        except (OSError, RuntimeError, ValueError) as error:
            QMessageBox.critical(self, "Calibration image error", str(error))
            return
        if not self._confirm_discard_unsaved_regression():
            return

        self._calibration_source_path = normalized_path
        self._calibration_source_image = source_image
        self._last_calibration_directory = str(Path(normalized_path).parent)
        self.origImg = normalized_path
        self._origPixmap = pixmap
        self.ui.labelOrigImg.setText("")
        self._scale_label(self.ui.labelOrigImg)
        self._clear_calibration_results()
        self.clear_active_formulas_requested.emit()
        self._set_active_worker_task(None)

    @Slot()
    def _start_linear_regression(self):
        if self._active_worker_task is not None:
            QMessageBox.warning(self, "Linear regression", "Another task is already running.")
            return
        if not self._calibration_source_path or self._calibration_source_image is None:
            QMessageBox.critical(
                self, "Linear regression error", "Import a calibration image first."
            )
            return
        if not self._confirm_discard_unsaved_regression():
            return
        calibration_path = Path(self._calibration_source_path)
        if not calibration_path.is_file():
            QMessageBox.critical(
                self,
                "Linear regression error",
                "The calibration image was not found:\n{}".format(calibration_path),
            )
            return
        repository_root = Path(__file__).resolve().parent.parent
        weight_path = (
            repository_root
            / "HT-Detector_Peng"
            / "weights"
            / "cuvette_Peng"
            / "yolov8n_train"
            / "weights"
            / "best.pt"
        )
        if not weight_path.is_file():
            QMessageBox.critical(
                self,
                "Linear regression error",
                "YOLO weight file was not found:\n{}".format(weight_path),
            )
            return
        try:
            load_effective_settings(apply_to_module=False)
        except Exception as error:
            QMessageBox.critical(
                self,
                "Linear regression error",
                "Interface settings could not be loaded: {}".format(error),
            )
            return

        self.ui.progressBar.setRange(0, 0)
        self._set_active_worker_task("regression")
        self.regression_requested.emit(
            self._calibration_source_path, str(weight_path)
        )

    @Slot()
    def _select_detection_image(self):
        if self._active_worker_task is not None:
            QMessageBox.warning(self, "Detection", "Another task is already running.")
            return
        if self._detection_dirty:
            answer = QMessageBox.question(
                self,
                "Unsaved detection result",
                "当前检测结果尚未保存，继续检测将丢弃该结果。",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )
            if answer != QMessageBox.Yes:
                return
        image_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select an image for detection",
            "",
            "Images (*.jpg *.jpeg *.png *.bmp *.tif *.tiff)",
        )
        if not image_path:
            return

        repository_root = Path(__file__).resolve().parent.parent
        weight_path = (
            repository_root
            / "HT-Detector_Peng"
            / "weights"
            / "cuvette_Peng"
            / "yolov8n_train"
            / "weights"
            / "best.pt"
        )
        if not weight_path.is_file():
            QMessageBox.critical(
                self,
                "Detection error",
                "YOLO weight file was not found:\n{}".format(weight_path),
            )
            return

        self.ui.progressBar.setRange(0, 100)
        self.ui.progressBar.setValue(0)
        self.ui.progressBar.setRange(0, 0)
        self._set_active_worker_task("detection")
        self.detection_requested.emit(image_path, str(weight_path))

    @Slot(object)
    def _on_detection_finished(self, payload):
        try:
            export_payload = self._validated_detection_export_payload(payload)
            image = export_payload["image"]
            pixmap = self._bgr_image_to_pixmap(image)
            if pixmap.isNull():
                raise ValueError("The detection annotated image could not be displayed.")
            headers = ["No.", "Con.", "Red", "Green", "Blue"]
            rows = [
                tuple(target[header] for header in headers)
                for target in export_payload["targets"]
            ]
            model = self._build_table_model(headers, rows)

            previous_result = self._detection_result
            previous_dirty = self._detection_dirty
            previous_type = self._last_completed_result_type
            previous_pixmap = self._recgPixmap
            previous_model = self.ui.tabviewRecg.model()
            try:
                self._recgPixmap = pixmap
                self._scale_label(self.ui.labelRecgImg)
                self.ui.tabviewRecg.setModel(model)
                self.ui.tabviewRecg.verticalHeader().hide()
                self.ui.tabviewRecg.horizontalHeader().setSectionResizeMode(
                    QHeaderView.Stretch
                )
                self._detection_result = payload
                self._detection_dirty = True
                self._last_completed_result_type = "detection"
            except Exception:
                self._detection_result = previous_result
                self._detection_dirty = previous_dirty
                self._last_completed_result_type = previous_type
                self._recgPixmap = previous_pixmap
                self.ui.tabviewRecg.setModel(previous_model)
                self._scale_label(self.ui.labelRecgImg)
                raise
            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(100)
            warnings = [str(message) for message in payload.get("warnings", []) if message]
            if warnings:
                self._show_message_safely(
                    QMessageBox.warning, self, "Detection warning", "\n".join(warnings)
                )
        except Exception as error:
            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(0)
            self._show_message_safely(
                QMessageBox.warning,
                self,
                "Detection warning",
                "The new detection result was not accepted; the previous result was preserved.\n{}".format(
                    error
                ),
            )
        finally:
            self._set_active_worker_task(None)

    @Slot(str)
    def _on_detection_failed(self, message):
        try:
            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(0)
            self._show_message_safely(
                QMessageBox.critical, self, "Detection error", message
            )
        finally:
            self._set_active_worker_task(None)

    @Slot(object)
    def _on_regression_finished(self, payload):
        previous_result = self._regression_result
        previous_dirty = self._regression_dirty
        previous_type = self._last_completed_result_type
        previous_orig_img = self.origImg
        previous_pixmap = self._origPixmap
        previous_model = self.ui.tabviewOrig.model()
        previous_plot_has_result = self._regression_plot_has_result
        previous_elapsed = self.ui.lcdNumber.value()
        try:
            export_payload = self._validated_linear_export_payload(payload)
            pixmap = self._bgr_image_to_pixmap(export_payload["image"])
            if pixmap.isNull():
                raise ValueError("The regression annotated image could not be displayed.")
            headers = list(self.CALIBRATION_TABLE_HEADERS)
            rows = [
                tuple(sample[header] for header in headers)
                for sample in export_payload["samples"]
            ]
            model = self._build_table_model(headers, rows)

            try:
                self.origImg = export_payload["image"]
                self._origPixmap = pixmap
                self._scale_label(self.ui.labelOrigImg)
                self.ui.tabviewOrig.setModel(model)
                self.ui.tabviewOrig.verticalHeader().hide()
                self.ui.tabviewOrig.horizontalHeader().setSectionResizeMode(
                    QHeaderView.Stretch
                )
                self._regression_result = payload
                self._regression_dirty = True
                self._last_completed_result_type = "linear"
                self._show_calibration_plot_placeholder()
                self.ui.lcdNumber.display(int(round(payload["elapsed_ms"])))
            except Exception:
                self._regression_result = previous_result
                self._regression_dirty = previous_dirty
                self._last_completed_result_type = previous_type
                self.origImg = previous_orig_img
                self._origPixmap = previous_pixmap
                self.ui.tabviewOrig.setModel(previous_model)
                self._scale_label(self.ui.labelOrigImg)
                self.ui.lcdNumber.display(previous_elapsed)
                if previous_result is not None and previous_plot_has_result:
                    self._plot_regression_result()
                else:
                    self._show_calibration_plot_placeholder()
                raise

            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(100)
            warnings = [str(message) for message in payload.get("warnings", []) if message]
            if warnings:
                self._show_message_safely(
                    QMessageBox.warning,
                    self,
                    "Linear regression warning",
                    "\n".join(warnings),
                )
        except Exception as error:
            self._restore_previous_regression_formulas(previous_result, previous_dirty)
            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(0)
            self._show_message_safely(
                QMessageBox.critical,
                self,
                "Linear regression error",
                "The new regression result was not accepted; the previous result was preserved.\n{}".format(
                    error
                ),
            )
        finally:
            self._set_active_worker_task(None)

    @Slot(str)
    def _on_regression_failed(self, message):
        try:
            self._restore_previous_regression_formulas(
                self._regression_result, self._regression_dirty
            )
            self.ui.progressBar.setRange(0, 100)
            self.ui.progressBar.setValue(0)
            self._show_message_safely(
                QMessageBox.critical, self, "Linear regression error", message
            )
        finally:
            self._set_active_worker_task(None)

    def _restore_previous_regression_formulas(self, result, dirty):
        if dirty and isinstance(result, dict):
            formulas = result.get("formulas")
            if isinstance(formulas, dict):
                self.restore_active_formulas_requested.emit(formulas)
                return
        self.clear_active_formulas_requested.emit()

    @staticmethod
    def _show_message_safely(message_function, *args):
        try:
            message_function(*args)
        except Exception:
            pass

    @Slot()
    def _shutdown_detection_thread(self):
        if self._detection_thread.isRunning():
            self._detection_thread.quit()
            self._detection_thread.wait()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_image_scales()

    def eventFilter(self, obj, event):
        if event.type() == QEvent.Type.Resize and obj in (
                getattr(self.ui, 'labelOrigImg', None),
                getattr(self.ui, 'labelRecgImg', None)):
            self._scale_label(obj)
        return super().eventFilter(obj, event)

    def _update_image_scales(self):
        for label in (getattr(self.ui, 'labelOrigImg', None),
                      getattr(self.ui, 'labelRecgImg', None)):
            if label is not None:
                self._scale_label(label)

    def _install_resizable_splitters(self):
        self._install_linear_regression_splitter()
        self._install_detection_splitter()

    @staticmethod
    def _configure_splitter(splitter, sizes):
        splitter.setChildrenCollapsible(False)
        splitter.setHandleWidth(6)
        for index, stretch in enumerate(sizes):
            splitter.setStretchFactor(index, stretch)
        splitter.setSizes([size * 100 for size in sizes])

    def _install_linear_regression_splitter(self):
        self.ui.horizontalLayout_6.removeWidget(self.ui.tabWidget_2)
        self.ui.verticalLayout_8.removeWidget(self.ui.label_2)
        self.ui.verticalLayout_8.removeWidget(self.ui.tabviewOrig)

        table_panel = QWidget(self.ui.cameraMainGroupBox)
        table_panel.setObjectName("linearRegressionTablePanel")
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(self.ui.label_2)
        table_layout.addWidget(self.ui.tabviewOrig)

        splitter = QSplitter(Qt.Horizontal, self.ui.cameraMainGroupBox)
        splitter.setObjectName("linearRegressionSplitter")
        splitter.addWidget(self.ui.tabWidget_2)
        splitter.addWidget(table_panel)
        self._configure_splitter(splitter, [5, 2])

        self.ui.horizontalLayout_9.removeItem(self.ui.horizontalLayout_6)
        self.ui.horizontalLayout_9.addWidget(splitter)

    def _install_detection_splitter(self):
        self.ui.verticalLayout_10.removeWidget(self.ui.tabWidget)
        for widget in (self.ui.label_5, self.ui.progressBar,
                       self.ui.label_9, self.ui.lcdNumber):
            self.ui.horizontalLayout_2.removeWidget(widget)
        self.ui.verticalLayout_9.removeWidget(self.ui.label_4)
        self.ui.verticalLayout_9.removeWidget(self.ui.tabviewRecg)

        image_panel = QWidget(self.ui.groupBox_3)
        image_panel.setObjectName("detectionImagePanel")
        image_layout = QVBoxLayout(image_panel)
        image_layout.setContentsMargins(0, 0, 0, 0)
        image_layout.addWidget(self.ui.tabWidget)
        progress_layout = QHBoxLayout()
        progress_layout.addWidget(self.ui.label_5)
        progress_layout.addWidget(self.ui.progressBar)
        progress_layout.addWidget(self.ui.label_9)
        progress_layout.addWidget(self.ui.lcdNumber)
        image_layout.addLayout(progress_layout)

        table_panel = QWidget(self.ui.groupBox_3)
        table_panel.setObjectName("detectionTablePanel")
        table_layout = QVBoxLayout(table_panel)
        table_layout.setContentsMargins(0, 0, 0, 0)
        table_layout.addWidget(self.ui.label_4)
        table_layout.addWidget(self.ui.tabviewRecg)

        splitter = QSplitter(Qt.Horizontal, self.ui.groupBox_3)
        splitter.setObjectName("detectionSplitter")
        splitter.addWidget(image_panel)
        splitter.addWidget(table_panel)
        self._configure_splitter(splitter, [5, 2])

        self.ui.verticalLayout_11.removeItem(self.ui.horizontalLayout_4)
        self.ui.verticalLayout_11.addWidget(splitter)

    def _scale_label(self, label):
        if label is getattr(self.ui, 'labelOrigImg', None):
            pix = getattr(self, '_origPixmap', None)
        elif label is getattr(self.ui, 'labelRecgImg', None):
            pix = getattr(self, '_recgPixmap', None)
        else:
            return
        if pix is None or pix.isNull():
            return
        lw, lh = label.width(), label.height()
        if lw <= 0 or lh <= 0:
            return
        # KeepAspectRatioByExpanding fills the target rectangle completely without
        # distortion; we then crop-center so the label never exceeds its own bounds.
        scaled = pix.scaled(lw, lh, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
        sw, sh = scaled.width(), scaled.height()
        if sw > lw or sh > lh:
            x = max(0, (sw - lw) // 2)
            y = max(0, (sh - lh) // 2)
            scaled = scaled.copy(x, y, min(lw, sw), min(lh, sh))
        label.setPixmap(scaled)

    @classmethod
    def _resolve_image(cls, base_no_ext):
        for ext in cls.IMAGE_EXTENSIONS:
            candidate = base_no_ext + ext
            if os.path.exists(candidate):
                return candidate
        return base_no_ext + cls.IMAGE_EXTENSIONS[0]

    @staticmethod
    def _read_excel(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h) for h in next(it)]
        rows = [r for r in it if any(c is not None for c in r)]
        return headers, rows

    @staticmethod
    def _first_color_index(headers):
        return headers.index("Con.") + 1

    @staticmethod
    def _build_table_model(headers, rows):
        model = QStandardItemModel()
        model.setColumnCount(len(headers))
        for c, h in enumerate(headers):
            model.setHeaderData(c, Qt.Horizontal, h)
        for r, row in enumerate(rows):
            for c, val in enumerate(row):
                if headers[c] == "No.":
                    text = str(int(val))
                elif isinstance(val, (int, float)):
                    text = "{:.2f}".format(round(float(val), 2))
                else:
                    text = "" if val is None else str(val)
                model.setItem(r, c, QStandardItem(text))
        return model

    @classmethod
    def _populate_tableview(cls, view, headers, rows):
        model = cls._build_table_model(headers, rows)
        view.setModel(model)
        view.verticalHeader().hide()
        for c in range(len(headers)):
            view.setColumnWidth(c, view.width() // 2)


# self.ui.labelOrigImg.width(),

#class Camera(QMainWindow):
#class ImageSettings(QDialog):
