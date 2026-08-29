import math
import numbers
import time
from pathlib import Path

from PySide6.QtCore import QObject, Signal, Slot

from interface_config import load_effective_settings


def _box_center(box):
    return ((box[0] + box[2]) / 2.0, (box[1] + box[3]) / 2.0)


def _intersection_area(first, second):
    width = max(0.0, min(first[2], second[2]) - max(first[0], second[0]))
    height = max(0.0, min(first[3], second[3]) - max(first[1], second[1]))
    return width * height


def _pair_cuvettes_and_liquids(cuvettes, liquids):
    cuvettes = sorted(cuvettes, key=lambda box: box[0])
    liquids = sorted(liquids, key=lambda box: box[0])
    unused_cuvettes = set(range(len(cuvettes)))
    pairs = []
    unmatched_liquids = []

    for liquid in liquids:
        liquid_center = _box_center(liquid)
        liquid_area = max(0.0, liquid[2] - liquid[0]) * max(0.0, liquid[3] - liquid[1])
        contained = []
        overlapping = []
        for index in unused_cuvettes:
            cuvette = cuvettes[index]
            cuvette_center = _box_center(cuvette)
            center_distance = math.hypot(
                liquid_center[0] - cuvette_center[0],
                liquid_center[1] - cuvette_center[1],
            )
            coverage = (
                _intersection_area(liquid, cuvette) / liquid_area
                if liquid_area > 0
                else 0.0
            )
            center_inside = (
                cuvette[0] <= liquid_center[0] <= cuvette[2]
                and cuvette[1] <= liquid_center[1] <= cuvette[3]
            )
            horizontal_overlap = min(liquid[2], cuvette[2]) - max(
                liquid[0], cuvette[0]
            )
            if center_inside:
                contained.append((-coverage, center_distance, index))
            elif horizontal_overlap > 0:
                overlapping.append((center_distance, -horizontal_overlap, index))

        candidates = contained if contained else overlapping
        if not candidates:
            unmatched_liquids.append(liquid)
            continue
        selected_index = min(candidates)[-1]
        unused_cuvettes.remove(selected_index)
        pairs.append((cuvettes[selected_index], liquid))

    pairs.sort(key=lambda pair: min(pair[0][0], pair[1][0]))
    unmatched_cuvettes = [cuvettes[index] for index in sorted(unused_cuvettes)]
    return pairs, unmatched_cuvettes, unmatched_liquids


def _calculate_rgb_roi(liquid_box, image_shape, ratios):
    x0, y0, x1, y1 = liquid_box
    x0_ratio, y0_ratio, x1_ratio, y1_ratio = ratios
    x0_con = int(x0 + (x1 - x0) * x0_ratio)
    y0_con = int(y0 + (y1 - y0) * y0_ratio)
    x1_con = int(x0 + (x1 - x0) * x1_ratio)
    y1_con = int(y0 + (y1 - y0) * y1_ratio)

    height, width = image_shape[:2]
    x0_con = min(max(x0_con, 0), width)
    x1_con = min(max(x1_con, 0), width)
    y0_con = min(max(y0_con, 0), height)
    y1_con = min(max(y1_con, 0), height)
    if x1_con <= x0_con or y1_con <= y0_con:
        raise ValueError("The RGB sampling region is empty after clipping.")
    return x0_con, y0_con, x1_con, y1_con


def _calculate_measurement(original_bgr, roi, accuracy, channel, slope, intercept):
    red, green, blue = _calculate_rgb_averages(original_bgr, roi, accuracy)
    values = {"R": red, "G": green, "B": blue}
    concentration = (values[channel] - intercept) / slope
    if not math.isfinite(concentration):
        raise ValueError("The calculated concentration is not finite.")
    return red, green, blue, concentration


def _calculate_rgb_averages(original_bgr, roi, accuracy):
    x0, y0, x1, y1 = roi
    region_bgr = original_bgr[y0:y1, x0:x1]
    if region_bgr.size == 0:
        raise ValueError("The RGB sampling region contains no pixels.")
    region_rgb = region_bgr[..., ::-1]
    averages = region_rgb.reshape(-1, 3).mean(axis=0)
    red, green, blue = (
        round(float(value), accuracy) for value in averages
    )
    if not all(math.isfinite(value) for value in (red, green, blue)):
        raise ValueError("The calculated RGB values are not finite.")
    return red, green, blue


def _validated_display_order(order):
    normalized = str(order).strip().upper()
    if normalized.startswith("CON") and len(normalized) == 6 and set(normalized[3:]) == {"R", "G", "B"}:
        return normalized, None
    return "CONRGB", "Invalid Order_Con_R_G_B; using ConRGB."


def _text_layout(cv2, image_height, texts):
    target_height = min(max(image_height * 0.03, 14.0), 64.0)
    thickness = min(max(int(round(target_height / 12.0)), 1), 5)
    scale = min(max(target_height / 22.0, 0.55), 3.0)
    probe_size = cv2.getTextSize(
        "Con.:000.00", cv2.FONT_HERSHEY_SIMPLEX, scale, thickness
    )[0]
    if probe_size[1] > 0:
        scale = min(max(scale * target_height / probe_size[1], 0.55), 3.0)

    sizes = [
        cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness)
        for text in texts
    ]
    max_width = max((size[0][0] for size in sizes), default=0)
    max_text_height = max((size[0][1] for size in sizes), default=int(target_height))
    max_baseline = max((size[1] for size in sizes), default=0)
    line_height = max_text_height + max_baseline + max(4, thickness * 2)
    return {
        "font_scale": scale,
        "thickness": thickness,
        "line_height": line_height,
        "block_width": max_width,
        "block_height": line_height * len(texts),
        "sizes": sizes,
    }


class YoloDetectionWorker(QObject):
    finished = Signal(object)
    failed = Signal(str)
    regression_finished = Signal(object)
    regression_failed = Signal(str)

    FORMULA_CHANNELS = ("R", "G", "B")
    FORMULA_FIELDS = ("slope", "intercept", "r", "R2", "p", "std_err")

    def __init__(self, parent=None):
        super().__init__(parent)
        self._model = None
        self._weight_path = None
        self._formula_cache = {}
        self._active_formulas = None

    @classmethod
    def _validated_formulas(cls, formulas, source="regression formulas"):
        if not isinstance(formulas, dict) or set(formulas) != set(cls.FORMULA_CHANNELS):
            raise ValueError("{} must contain exactly R, G, and B.".format(source))
        validated = {}
        for channel in cls.FORMULA_CHANNELS:
            formula = formulas[channel]
            if not isinstance(formula, dict):
                raise ValueError("{} channel {} is invalid.".format(source, channel))
            values = {}
            for field in cls.FORMULA_FIELDS:
                value = formula.get(field)
                if isinstance(value, bool) or not isinstance(value, numbers.Real):
                    raise ValueError(
                        "{} {}.{} must be numeric.".format(source, channel, field)
                    )
                value = float(value)
                if not math.isfinite(value):
                    raise ValueError(
                        "{} {}.{} must be finite.".format(source, channel, field)
                    )
                values[field] = value
            if math.isclose(values["slope"], 0.0, rel_tol=0.0, abs_tol=1e-12):
                raise ValueError("{} {} slope must not be zero or near zero.".format(source, channel))
            validated[channel] = values
        return validated

    @classmethod
    def _read_merged_formulas(cls, formula_path):
        from openpyxl import load_workbook

        try:
            workbook = load_workbook(formula_path, read_only=True, data_only=True)
        except Exception as error:
            raise RuntimeError(
                "The merged standard curve file could not be opened: {}: {}".format(
                    formula_path, error
                )
            ) from error
        try:
            if "Sheet1" not in workbook.sheetnames:
                raise RuntimeError("Sheet1 is missing from: {}".format(formula_path))
            worksheet = workbook["Sheet1"]
            expected_headers = ("Channel",) + cls.FORMULA_FIELDS
            headers = tuple(worksheet.cell(1, column).value for column in range(8, 15))
            if headers != expected_headers:
                raise RuntimeError(
                    "The merged standard curve headers H1:N1 are invalid in: {}".format(
                        formula_path
                    )
                )
            formulas = {}
            for row in range(2, worksheet.max_row + 1):
                channel = worksheet.cell(row, 8).value
                values = tuple(worksheet.cell(row, column).value for column in range(9, 15))
                if channel is None and all(value is None for value in values):
                    continue
                if channel not in cls.FORMULA_CHANNELS:
                    raise RuntimeError(
                        "Invalid formula channel at H{} in: {}".format(row, formula_path)
                    )
                if channel in formulas:
                    raise RuntimeError(
                        "Duplicate formula channel {} in: {}".format(channel, formula_path)
                    )
                formulas[channel] = dict(zip(cls.FORMULA_FIELDS, values))
        finally:
            workbook.close()
        try:
            return cls._validated_formulas(formulas, str(formula_path))
        except ValueError as error:
            raise RuntimeError(str(error)) from error

    @staticmethod
    def _read_legacy_formula(formula_path):
        from openpyxl import load_workbook

        if not formula_path.is_file():
            raise RuntimeError("The standard curve file was not found: {}".format(formula_path))

        workbook = load_workbook(formula_path, read_only=True, data_only=True)
        try:
            if "Sheet1" not in workbook.sheetnames:
                raise RuntimeError("Sheet1 is missing from: {}".format(formula_path))
            worksheet = workbook["Sheet1"]
            if worksheet["A1"].value != "slope" or worksheet["B1"].value != "intercept":
                raise RuntimeError(
                    "The standard curve headers must be A1=slope and B1=intercept: {}".format(
                        formula_path
                    )
                )
            slope = worksheet["A2"].value
            intercept = worksheet["B2"].value
        finally:
            workbook.close()

        for name, value in (("slope", slope), ("intercept", intercept)):
            if isinstance(value, bool) or not isinstance(value, numbers.Real):
                raise RuntimeError("{} must be numeric in: {}".format(name, formula_path))
            if not math.isfinite(float(value)):
                raise RuntimeError("{} must be finite in: {}".format(name, formula_path))
        slope = float(slope)
        intercept = float(intercept)
        if math.isclose(slope, 0.0, rel_tol=0.0, abs_tol=1e-12):
            raise RuntimeError("slope must not be zero or near zero in: {}".format(formula_path))

        return slope, intercept

    def _load_formula(self, color_channel):
        if self._active_formulas is not None:
            formula = self._active_formulas[color_channel]
            return formula["slope"], formula["intercept"]
        if color_channel in self._formula_cache:
            return self._formula_cache[color_channel]

        repository_root = Path(__file__).resolve().parent.parent
        results_root = repository_root / "HT-Detector_Peng" / "runs" / "detect" / "results"
        merged_path = results_root / "linear" / "linear_con_rgb.xlsx"
        if merged_path.is_file():
            formulas = self._read_merged_formulas(merged_path)
            self._formula_cache.update(
                {
                    channel: (formula["slope"], formula["intercept"])
                    for channel, formula in formulas.items()
                }
            )
            return self._formula_cache[color_channel]

        legacy_path = results_root / "linear_formula_{}.xlsx".format(color_channel)
        formula = self._read_legacy_formula(legacy_path)
        self._formula_cache[color_channel] = formula
        return formula

    def _get_model(self, weight_path):
        if self._model is None or self._weight_path != weight_path:
            from ultralytics import YOLO

            self._model = YOLO(weight_path)
            self._weight_path = weight_path
        return self._model

    @Slot()
    def clear_active_formulas(self):
        self._active_formulas = None

    @Slot(object)
    def install_saved_formulas(self, formulas):
        validated = self._validated_formulas(formulas, "saved regression formulas")
        self._active_formulas = validated
        self._formula_cache = {
            channel: (formula["slope"], formula["intercept"])
            for channel, formula in validated.items()
        }

    @staticmethod
    def _class_ids(names):
        items = names.items() if isinstance(names, dict) else enumerate(names)
        normalized = {str(name).strip().lower(): int(class_id) for class_id, name in items}
        missing = [name for name in ("cuvette", "liquid") if name not in normalized]
        if missing:
            raise RuntimeError(
                "Model class names are missing: {}.".format(", ".join(missing))
            )
        return normalized["cuvette"], normalized["liquid"]

    @staticmethod
    def _put_text(
        cv2, image, text, x, y, color, scale=0.55, thickness=1, outline=False
    ):
        height, width = image.shape[:2]
        (text_width, text_height), baseline = cv2.getTextSize(
            text, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness
        )
        x = min(max(int(x), 0), max(0, width - text_width - 1))
        y = min(max(int(y), text_height + baseline), max(text_height + baseline, height - baseline - 1))
        if outline:
            cv2.putText(
                image,
                text,
                (x, y),
                cv2.FONT_HERSHEY_SIMPLEX,
                scale,
                (0, 0, 0),
                thickness + 2,
                cv2.LINE_AA,
            )
        cv2.putText(
            image,
            text,
            (x, y),
            cv2.FONT_HERSHEY_SIMPLEX,
            scale,
            color,
            thickness,
            cv2.LINE_AA,
        )

    @classmethod
    def _draw_detection_label(cls, cv2, image, text, box_x0, box_y0, color):
        image_height, image_width = image.shape[:2]
        target_height = min(max(image_height * 0.025, 14.0), 64.0)
        thickness = min(max(int(round(target_height / 12.0)), 1), 5)
        probe = "cuvette 0.00"
        probe_height = cv2.getTextSize(
            probe, cv2.FONT_HERSHEY_SIMPLEX, 1.0, thickness
        )[0][1]
        scale = target_height / max(probe_height, 1)
        calibrated_height = cv2.getTextSize(
            probe, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness
        )[0][1]
        if calibrated_height > 0:
            scale *= target_height / calibrated_height

        (text_width, text_height), baseline = cv2.getTextSize(
            text, cv2.FONT_HERSHEY_SIMPLEX, scale, thickness
        )
        outline_padding = thickness + 2
        gap = max(2, thickness)
        max_x = max(outline_padding, image_width - text_width - outline_padding)
        x = min(max(int(box_x0), outline_padding), max_x)

        outside_y = int(box_y0) - gap - baseline - outline_padding
        if outside_y - text_height - outline_padding >= 0:
            y = outside_y
        else:
            y = int(box_y0) + gap + outline_padding + text_height
        min_y = text_height + outline_padding
        max_y = max(min_y, image_height - baseline - outline_padding)
        y = min(max(y, min_y), max_y)

        cls._put_text(
            cv2,
            image,
            text,
            x,
            y,
            color,
            scale=scale,
            thickness=thickness,
            outline=True,
        )

    @classmethod
    def _draw_text_block(cls, cv2, image, text_lines, anchor_x, preferred_top):
        texts = [text for text, _ in text_lines]
        layout = _text_layout(cv2, image.shape[0], texts)
        image_height, image_width = image.shape[:2]
        outline_padding = layout["thickness"] + 2
        block_width = layout["block_width"] + outline_padding * 2
        block_height = layout["block_height"] + outline_padding * 2
        start_x = min(
            max(int(anchor_x), outline_padding),
            max(outline_padding, image_width - block_width),
        )
        start_top = min(
            max(int(preferred_top), outline_padding),
            max(outline_padding, image_height - block_height),
        )

        for index, (text, color) in enumerate(text_lines):
            (_, text_height), baseline = layout["sizes"][index]
            baseline_y = (
                start_top
                + index * layout["line_height"]
                + text_height
                + baseline
            )
            cls._put_text(
                cv2,
                image,
                text,
                start_x,
                baseline_y,
                color,
                scale=layout["font_scale"],
                thickness=layout["thickness"],
                outline=True,
            )
        return start_x, start_top, layout

    def _build_payload(self, result, settings, cv2, config_warnings=None):
        import numpy as np

        original_bgr = result.orig_img
        if not isinstance(original_bgr, np.ndarray) or original_bgr.ndim != 3:
            raise RuntimeError("YOLO did not return a valid OpenCV source image.")
        annotated = original_bgr.copy()
        warnings = list(config_warnings or [])
        targets = []

        color_channel = settings["color_channel"]
        if color_channel not in {"R", "G", "B"}:
            raise RuntimeError("color_channel must be R, G, or B.")
        display_order, order_warning = _validated_display_order(
            settings["Order_Con_R_G_B"]
        )
        if order_warning:
            warnings.append(order_warning)

        boxes = result.boxes
        if boxes is None or len(boxes) == 0:
            warnings.append("Detection completed, but no objects were found.")
            return {"image": annotated, "targets": targets, "warnings": warnings}

        coordinates = boxes.xyxy.detach().cpu().numpy()
        classes = boxes.cls.detach().cpu().numpy().astype(int)
        confidences = boxes.conf.detach().cpu().numpy()
        cuvette_id, liquid_id = self._class_ids(result.names)
        cuvettes = []
        liquids = []

        for coordinates_xyxy, class_id, confidence in zip(
            coordinates, classes, confidences
        ):
            box = tuple(float(value) for value in coordinates_xyxy)
            if class_id == cuvette_id:
                cuvettes.append(box)
            elif class_id == liquid_id:
                liquids.append(box)

            x0, y0, x1, y1 = (int(value) for value in box)
            x0 = min(max(x0, 0), annotated.shape[1] - 1)
            x1 = min(max(x1, 0), annotated.shape[1] - 1)
            y0 = min(max(y0, 0), annotated.shape[0] - 1)
            y1 = min(max(y1, 0), annotated.shape[0] - 1)
            cv2.rectangle(annotated, (x0, y0), (x1, y1), (0, 255, 0), 2)
            class_name = result.names[class_id]
            label = str(class_name)
            if settings["show_confidence"]:
                label += " {:.2f}".format(float(confidence))
            self._draw_detection_label(
                cv2, annotated, label, x0, y0, (0, 255, 0)
            )

        pairs, unmatched_cuvettes, unmatched_liquids = _pair_cuvettes_and_liquids(
            cuvettes, liquids
        )
        if unmatched_cuvettes:
            warnings.append("{} cuvette box(es) were not matched.".format(len(unmatched_cuvettes)))
        if unmatched_liquids:
            warnings.append("{} liquid box(es) were not matched.".format(len(unmatched_liquids)))
        if not pairs:
            warnings.append("No valid cuvette-liquid pairs were found.")
            return {"image": annotated, "targets": targets, "warnings": warnings}

        slope, intercept = self._load_formula(color_channel)
        ratios = (
            settings["x0_ratio"],
            settings["y0_ratio"],
            settings["x1_ratio"],
            settings["y1_ratio"],
        )
        rgb_colors = {"R": (0, 0, 255), "G": (0, 255, 0), "B": (255, 0, 0)}
        for number, (cuvette_box, liquid_box) in enumerate(pairs, start=1):
            try:
                roi = _calculate_rgb_roi(liquid_box, original_bgr.shape, ratios)
                red, green, blue, concentration = _calculate_measurement(
                    original_bgr,
                    roi,
                    settings["rgb_calculate_accuracy"],
                    color_channel,
                    slope,
                    intercept,
                )
            except (TypeError, ValueError) as error:
                warnings.append("No.{} was skipped: {}".format(number, error))
                continue

            target_number = len(targets) + 1
            targets.append(
                {
                    "No.": target_number,
                    "Con.": concentration,
                    "Red": red,
                    "Green": green,
                    "Blue": blue,
                    "cuvette_box": cuvette_box,
                    "liquid_box": liquid_box,
                    "rgb_roi": roi,
                }
            )
            x0_roi, y0_roi, x1_roi, y1_roi = roi
            cv2.rectangle(
                annotated, (x0_roi, y0_roi), (x1_roi, y1_roi), (250, 240, 10), 2
            )

            text_values = {
                "R": round(red, settings["rgb_display_accuracy"]),
                "G": round(green, settings["rgb_display_accuracy"]),
                "B": round(blue, settings["rgb_display_accuracy"]),
            }
            con_text = round(concentration, settings["con_display_accuracy"])
            text_lines = [("No.{}".format(target_number), (255, 0, 255)),
                          ("Con.:{}".format(con_text), (255, 255, 0))]
            text_lines.extend(
                ("{}:{}".format(channel, text_values[channel]), rgb_colors[channel])
                for channel in display_order[3:]
            )
            anchor_x = int(cuvette_box[0])
            layout = _text_layout(
                cv2, annotated.shape[0], [text for text, _ in text_lines]
            )
            preferred_top = int(cuvette_box[3]) + layout["thickness"] + 4
            if preferred_top + layout["block_height"] > annotated.shape[0]:
                preferred_top = int(cuvette_box[1]) - layout["block_height"] - 4
            self._draw_text_block(
                cv2, annotated, text_lines, anchor_x, preferred_top
            )

        if not targets:
            warnings.append("No valid RGB measurements were produced.")
        return {"image": annotated, "targets": targets, "warnings": warnings}

    @staticmethod
    def _regression_formula(concentrations, values, channel):
        import numpy as np
        from scipy import stats

        if len(concentrations) < 2:
            raise ValueError("At least two included calibration points are required.")
        if len(set(concentrations)) < 2:
            raise ValueError("Included calibration concentrations must contain at least two distinct values.")
        if not all(math.isfinite(value) for value in concentrations + values):
            raise ValueError("{} regression inputs must all be finite.".format(channel))
        try:
            result = stats.linregress(concentrations, values)
        except Exception as error:
            raise ValueError("{} linear regression failed: {}".format(channel, error)) from error
        formula = {
            "slope": float(result.slope),
            "intercept": float(result.intercept),
            "r": float(result.rvalue),
            "R2": float(result.rvalue * result.rvalue),
            "p": float(result.pvalue),
            "std_err": float(result.stderr),
        }
        if not all(np.isfinite(value) for value in formula.values()):
            raise ValueError("{} linear regression produced a non-finite result.".format(channel))
        if math.isclose(formula["slope"], 0.0, rel_tol=0.0, abs_tol=1e-12):
            raise ValueError(
                "{} linear regression produced a zero or near-zero slope.".format(
                    channel
                )
            )
        return formula

    def _build_regression_payload(
        self, result, settings, cv2, source_path="", config_warnings=None
    ):
        import numpy as np

        original_bgr = result.orig_img
        if not isinstance(original_bgr, np.ndarray) or original_bgr.ndim != 3:
            raise RuntimeError("YOLO did not return a valid OpenCV source image.")
        annotated = original_bgr.copy()
        warnings = list(config_warnings or [])
        display_order, order_warning = _validated_display_order(
            settings["Order_Con_R_G_B"]
        )
        if order_warning:
            warnings.append(order_warning)

        boxes = result.boxes
        if boxes is None or len(boxes) == 0:
            raise ValueError("Calibration failed: no cuvette or liquid boxes were detected.")
        coordinates = boxes.xyxy.detach().cpu().numpy()
        classes = boxes.cls.detach().cpu().numpy().astype(int)
        confidences = boxes.conf.detach().cpu().numpy()
        cuvette_id, liquid_id = self._class_ids(result.names)
        cuvettes = []
        liquids = []
        for coordinates_xyxy, class_id, confidence in zip(
            coordinates, classes, confidences
        ):
            box = tuple(float(value) for value in coordinates_xyxy)
            if class_id == cuvette_id:
                cuvettes.append(box)
            elif class_id == liquid_id:
                liquids.append(box)

            x0, y0, x1, y1 = (int(value) for value in box)
            x0 = min(max(x0, 0), annotated.shape[1] - 1)
            x1 = min(max(x1, 0), annotated.shape[1] - 1)
            y0 = min(max(y0, 0), annotated.shape[0] - 1)
            y1 = min(max(y1, 0), annotated.shape[0] - 1)
            cv2.rectangle(annotated, (x0, y0), (x1, y1), (0, 255, 0), 2)
            label = str(result.names[class_id])
            if settings["show_confidence"]:
                label += " {:.2f}".format(float(confidence))
            self._draw_detection_label(cv2, annotated, label, x0, y0, (0, 255, 0))

        if not cuvettes:
            raise ValueError("Calibration failed: no cuvette boxes were detected.")
        if not liquids:
            raise ValueError("Calibration failed: no liquid boxes were detected.")
        pairs, unmatched_cuvettes, unmatched_liquids = _pair_cuvettes_and_liquids(
            cuvettes, liquids
        )
        concentration_count = len(settings["con_list"])
        if unmatched_cuvettes or unmatched_liquids or len(pairs) != concentration_count:
            raise ValueError(
                "Calibration pairing mismatch: {} valid pair(s), {} configured concentration(s), "
                "{} unmatched cuvette box(es), {} unmatched liquid box(es).".format(
                    len(pairs),
                    concentration_count,
                    len(unmatched_cuvettes),
                    len(unmatched_liquids),
                )
            )

        ratios = (
            settings["x0_ratio"],
            settings["y0_ratio"],
            settings["x1_ratio"],
            settings["y1_ratio"],
        )
        rgb_colors = {"R": (0, 0, 255), "G": (0, 255, 0), "B": (255, 0, 0)}
        samples = []
        for index, (cuvette_box, liquid_box) in enumerate(pairs):
            try:
                roi = _calculate_rgb_roi(liquid_box, original_bgr.shape, ratios)
                red, green, blue = _calculate_rgb_averages(
                    original_bgr, roi, settings["rgb_calculate_accuracy"]
                )
            except (TypeError, ValueError) as error:
                raise ValueError("Calibration sample No.{} failed: {}".format(index + 1, error)) from error
            concentration = float(settings["con_list"][index])
            included = bool(settings["linear_formula_point_matrix"][index])
            sample = {
                "No.": index + 1,
                "Con.": concentration,
                "Red": red,
                "Green": green,
                "Blue": blue,
                "included": included,
                "cuvette_box": cuvette_box,
                "liquid_box": liquid_box,
                "rgb_roi": roi,
            }
            samples.append(sample)
            x0_roi, y0_roi, x1_roi, y1_roi = roi
            cv2.rectangle(annotated, (x0_roi, y0_roi), (x1_roi, y1_roi), (250, 240, 10), 2)
            text_values = {
                "R": round(red, settings["rgb_display_accuracy"]),
                "G": round(green, settings["rgb_display_accuracy"]),
                "B": round(blue, settings["rgb_display_accuracy"]),
            }
            con_text = round(concentration, settings["con_display_accuracy"])
            text_lines = [
                ("No.{}".format(index + 1), (255, 0, 255)),
                ("Con.:{}".format(con_text), (255, 255, 0)),
            ]
            text_lines.extend(
                ("{}:{}".format(channel, text_values[channel]), rgb_colors[channel])
                for channel in display_order[3:]
            )
            anchor_x = int(cuvette_box[0])
            layout = _text_layout(cv2, annotated.shape[0], [text for text, _ in text_lines])
            preferred_top = int(cuvette_box[3]) + layout["thickness"] + 4
            if preferred_top + layout["block_height"] > annotated.shape[0]:
                preferred_top = int(cuvette_box[1]) - layout["block_height"] - 4
            self._draw_text_block(cv2, annotated, text_lines, anchor_x, preferred_top)

        included_samples = [sample for sample in samples if sample["included"]]
        if len(included_samples) < 2:
            raise ValueError("At least two calibration points must be included in regression.")
        formula_concentrations = [sample["Con."] for sample in included_samples]
        formulas = {
            "R": self._regression_formula(
                formula_concentrations, [sample["Red"] for sample in included_samples], "R"
            ),
            "G": self._regression_formula(
                formula_concentrations, [sample["Green"] for sample in included_samples], "G"
            ),
            "B": self._regression_formula(
                formula_concentrations, [sample["Blue"] for sample in included_samples], "B"
            ),
        }
        return {
            "source_path": str(source_path),
            "image": annotated,
            "samples": samples,
            "formulas": formulas,
            "selected_channel": settings["color_channel"],
            "warnings": warnings,
        }

    @Slot(str, str)
    def detect(self, image_path, weight_path):
        try:
            import cv2
            import numpy as np

            settings, config_warnings, _ = load_effective_settings()
            encoded_path = np.fromfile(image_path, dtype=np.uint8)
            source_image = cv2.imdecode(encoded_path, cv2.IMREAD_COLOR)
            if source_image is None:
                raise ValueError("The selected image could not be read: {}".format(image_path))

            model = self._get_model(weight_path)
            results = model.predict(
                source=source_image,
                device="cpu",
                conf=settings["detect_confidence"],
                save=False,
                verbose=False,
            )
            if not results:
                raise RuntimeError("YOLO returned no result for the selected image.")

            payload = self._build_payload(
                results[0], settings, cv2, config_warnings=config_warnings
            )
            payload["source_path"] = str(image_path)
            self.finished.emit(payload)
        except Exception as error:
            self.failed.emit("{}: {}".format(type(error).__name__, error))

    @Slot(str, str)
    def regress(self, image_path, weight_path):
        started = time.perf_counter()
        self._active_formulas = None
        try:
            import cv2
            import numpy as np

            settings, config_warnings, _ = load_effective_settings()
            encoded_path = np.fromfile(image_path, dtype=np.uint8)
            if encoded_path.size == 0:
                raise ValueError("The calibration image is empty: {}".format(image_path))
            source_image = cv2.imdecode(encoded_path, cv2.IMREAD_COLOR)
            if source_image is None or source_image.shape[0] <= 0 or source_image.shape[1] <= 0:
                raise ValueError("The calibration image could not be read: {}".format(image_path))

            model = self._get_model(weight_path)
            results = model.predict(
                source=source_image,
                device="cpu",
                conf=settings["detect_confidence"],
                save=False,
                verbose=False,
            )
            if not results:
                raise RuntimeError("YOLO returned no result for the calibration image.")
            payload = self._build_regression_payload(
                results[0],
                settings,
                cv2,
                source_path=image_path,
                config_warnings=config_warnings,
            )
            payload["elapsed_ms"] = (time.perf_counter() - started) * 1000.0
            self._active_formulas = payload["formulas"]
            self.regression_finished.emit(payload)
        except Exception as error:
            self._active_formulas = None
            self.regression_failed.emit("{}: {}".format(type(error).__name__, error))
